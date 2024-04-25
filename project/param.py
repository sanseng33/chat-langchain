


import pandas as pd
import os

from openpyxl import load_workbook


def append_or_create(sheet_name, row):
    file_path = sheet_name
    try:
        if not os.path.exists(file_path):
            # 如果文件不存在，直接创建并写入行
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                row.to_frame().T.to_excel(writer, index=False)
        else:
            # 如果文件存在，使用openpyxl加载工作簿进行追加
            book = load_workbook(file_path)
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                # 获取最后一个sheet的名称
                last_sheet_name = book.sheetnames[-1]
                startrow = book[last_sheet_name].max_row

                # 追加数据到新的一行
                row.to_frame().T.to_excel(writer, sheet_name=last_sheet_name, index=False, header=False,
                                          startrow=startrow)
    except Exception as e:
        print(f"处理行时出现异常，已跳过。异常信息: {row}")

# 读取xlsx文件
df = pd.read_excel('d://project_prompt.xlsx')  # 请将'原文件路径.xlsx'替换为你的文件路径

# 遍历DataFrame的每一行
for index, row in df.iterrows():
    try:
        # 检查第一列的内容是否包含'1'
        if '并且存在一个专利申请量各年份的趋势数据为' in str(row[0]):
            # 如果包含'1'，则将这一行写入到'1.xlsx'中
            append_or_create('d://data//技术背景.xlsx', row)
        # 检查第一列的内容是否包含'2'
        elif '得出以下技术现状分析的结论' in str(row[0]):
            # 如果包含'2'，则将这一行写入到'2.xlsx'中
            append_or_create('d://data//技术现状分析.xlsx', row)
        elif '技术预研报告中的“研究内容”部分' in str(row[0]):
            # 如果包含'2'，则将这一行写入到'2.xlsx'中
            append_or_create('d://data//研究内容.xlsx', row)
        elif '重点玩家分析' in str(row[0]):
            # 如果包含'2'，则将这一行写入到'2.xlsx'中
            append_or_create('d://data//重点玩家分析.xlsx', row)
        elif '参考IDC技术市场分类' in str(row[0]):
            # 如果包含'2'，则将这一行写入到'2.xlsx'中
            append_or_create('d://data//技术路线.xlsx', row)
        elif '结合行业技术体系和你自身的经验' in str(row[0]):
            # 如果包含'2'，则将这一行写入到'2.xlsx'中
            append_or_create('d://data//技术路线图表.xlsx', row)
        elif '对专利说明书的背景技术' in str(row[0]):
            # 如果包含'2'，则将这一行写入到'2.xlsx'中
            append_or_create('d://data//关键专利.xlsx', row)
    except Exception as e:
        print(f"处理行{index}时出现异常，已跳过。异常信息: {e}")